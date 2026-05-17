#!/usr/bin/env python3
import argparse
import json
from pathlib import Path


KEY_HINTS = {
    "pin",
    "port",
    "ball",
    "package",
    "function",
    "alternate",
    "register",
    "address",
    "bit",
    "reset",
    "access",
    "net",
    "signal",
    "connector",
    "module",
    "parameter",
}


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def guess_header_with_index(ws, max_scan_rows: int = 20) -> tuple[int | None, list[str]]:
    best_idx = None
    best_values: list[str] = []
    best_score = 0
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_scan_rows, ws.max_row), values_only=True), start=1):
        values = [cell_text(value) for value in row]
        nonempty = [value for value in values if value]
        score = len(nonempty)
        if any(any(hint in value.lower() for hint in KEY_HINTS) for value in nonempty):
            score += 5
        if score > best_score and len(nonempty) >= 2:
            best_score = score
            best_values = nonempty
            best_idx = idx
    return best_idx, best_values


def key_columns(headers: list[str]) -> list[str]:
    found = []
    for header in headers:
        lower = header.lower()
        if any(hint in lower for hint in KEY_HINTS):
            found.append(header)
    return found


def main() -> int:
    parser = argparse.ArgumentParser(description="Catalog an .xlsx/.xlsm workbook for embedded document indexing.")
    parser.add_argument("workbook", help="Workbook path")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl is required for catalog_excel.py")

    path = Path(args.workbook).resolve()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        header_row, headers = guess_header_with_index(ws)
        sheets.append(
            {
                "name": ws.title,
                "state": ws.sheet_state,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "header_guess": header_row,
                "key_columns": key_columns(headers),
            }
        )

    print(
        json.dumps(
            {
                "source_path": str(path).replace("\\", "/"),
                "workbook_type": path.suffix.lower().lstrip("."),
                "sheets": sheets,
            },
            indent=2,
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
